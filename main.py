import os
import logging
import asyncio
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from telegram.ext import (
    ApplicationBuilder, CommandHandler, CallbackQueryHandler,
    MessageHandler, ContextTypes, filters, ConversationHandler
)
from dotenv import load_dotenv
from openpyxl import load_workbook
import smtplib
from email.message import EmailMessage
import shutil
from datetime import datetime, date, timedelta
import calendar
import nest_asyncio

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)


# Загрузка переменных окружения
load_dotenv()

# Переменные окружения для бота и почты
BOT_TOKEN = os.getenv("BOT_TOKEN")
EMAIL_LOGIN = os.getenv("EMAIL_LOGIN")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
SMTP_SERVER = os.getenv("SMTP_SERVER")
SMTP_PORT = int(os.getenv("SMTP_PORT"))

# Новые переменные для адресов специалистов и копии
SPECIALIST_EMAIL_ALESYA = "bas@vds.by"
SPECIALIST_EMAIL_DMITRY = "bas2@vds.by"
CC_EMAIL = "bas@vds.by" # Всегда в копии

TEMPLATE_PATH = "template.xlsx" # Убедитесь, что template.xlsx существует в той же директории

# Состояния для ConversationHandler
PROJECT, OBJECT, NAME, UNIT, QUANTITY, MODULE, POSITION_DELIVERY_DATE, \
ATTACHMENT_CHOICE, FILE_INPUT, LINK_INPUT, \
CONFIRM_ADD_MORE, \
EDIT_MENU, SELECT_POSITION, EDIT_FIELD_SELECTION, EDIT_FIELD_INPUT, \
FINAL_CONFIRMATION, GLOBAL_DELIVERY_DATE_SELECTION, \
EDITING_UNIT, EDITING_MODULE, SPECIALIST_SELECTION = range(20) # Обновлено количество состояний

# Глобальные переменные для хранения данных пользователя и предварительно определенных списков
user_state = {}
projects = ["Stadler", "Мотели"]
objects = ["Мерке", "Аральск", "Атырау", "Каркаролинск", "Семипалатинск"]
modules = [f"{i+1}" for i in range(18)]
units = ["м", "м2", "м3", "шт", "компл", "л", "кг", "тн"]

def fill_excel(project, object_name, positions, user_full_name, telegram_id_or_username):
    """
    Заполняет Excel-файл данными, включая дату поставки для каждой позиции, проект, объект,
    а также информацию о пользователе, от которого пришла заявка.
    """
    today = datetime.today().strftime("%d.%m.%Y")
    sanitized_user_name = user_full_name.replace(" ", "_")
    filename = f"Заявка_{project}_{object_name}_{sanitized_user_name}_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
    output_dir = "out"

    template_full_path = os.path.abspath(TEMPLATE_PATH)

    os.makedirs(output_dir, exist_ok=True)
    new_path = os.path.join(output_dir, filename)

    shutil.copy(template_full_path, new_path)
    wb = load_workbook(new_path)
    ws = wb.active

    ws['G2'] = today
    ws['G3'] = project
    ws['G4'] = object_name
    ws['E16'] = user_full_name
    ws['E17'] = telegram_id_or_username

    logger.info(f"Writing to Excel: G2={today}, G3={project}, G4={object_name}, E16={user_full_name}, E17={telegram_id_or_username}")


    row_start_data = 9
    for i, pos in enumerate(positions):
        row = row_start_data + i

        ws.cell(row=row, column=1).value = i + 1
        ws.cell(row=row, column=2).value = pos["name"]
        ws.cell(row=row, column=3).value = pos["unit"]
        ws.cell(row=row, column=4).value = pos["quantity"]
        ws.cell(row=row, column=5).value = pos.get("delivery_date", "Не указано")
        ws.cell(row=row, column=6).value = pos["module"]
        ws.cell(row=row, column=7).value = pos.get("link", "")
        logger.info(f"Writing position {i+1} to Excel: {pos}")


    wb.save(new_path)
    logger.info(f"Excel file saved to: {new_path}")
    return new_path

async def send_email(chat_id, project, object_name, positions, user_full_name, telegram_id_or_username, to_email, cc_email, context=None):
    """
    Отправляет сгенерированный Excel-файл по электронной почте,
    с возможностью прикрепления дополнительных файлов и ссылок, привязанных к позициям,
    а также информацией о пользователе.
    """
    msg = EmailMessage()
    msg["Subject"] = f"Заявка на снабжение: {project} - {object_name}"
    msg["From"] = EMAIL_LOGIN
    msg["To"] = to_email
    msg["Cc"] = cc_email # Добавляем копию

    email_body = "Во вложении заявка на снабжение.\n\n"
    email_body += f"Проект: {project}\n"
    email_body += f"Объект: {object_name}\n"
    email_body += f"От кого: {user_full_name}\n"
    email_body += f"Telegram ID: {telegram_id_or_username}\n\n"
    email_body += "Позиции:\n"

    files_to_attach = []
    links_in_email = []

    for i, p in enumerate(positions):
        pos_info = (
            f"{i+1}. Модуль: {p.get('module', 'N/A')} | Наименование: {p.get('name', 'N/A')} | "
            f"Ед.изм.: {p.get('unit', 'N/A')} | Количество: {p.get('quantity', 'N/A')} | "
            f"Дата поставки: {p.get('delivery_date', 'N/A')}"
        )
        if p.get('link'):
            pos_info += f" | Ссылка: {p['link']}"
            links_in_email.append(f"Позиция {i+1} ({p.get('name', 'N/A')}): {p['link']}")

        if p.get('file_data') and isinstance(p['file_data'], list):
            file_names = []
            for file_item in p['file_data']:
                file_names.append(file_item.get('file_name', 'N/A'))
                files_to_attach.append((i+1, file_item))
            if file_names:
                pos_info += f" | Файлы: {', '.join(file_names)}"
        email_body += pos_info + "\n"

    if links_in_email:
        email_body += "\nОтдельные ссылки для позиций:\n" + "\n".join(links_in_email) + "\n"

    msg.set_content(email_body)
    logger.info(f"Email body generated for chat_id {chat_id}: \n{email_body}")

    try:
        file_path = fill_excel(project, object_name, positions, user_full_name, telegram_id_or_username)
        with open(file_path, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=os.path.basename(file_path),
            )
        logger.info(f"Excel file '{file_path}' прикреплен к письму")
    except Exception as e:
        logger.error(f"Ошибка при создании или прикреплении Excel файла: {e}")

    if context:
        for pos_index, file_data in files_to_attach:
            try:
                file_id = file_data['file_id']
                file_name = file_data['file_name']
                mime_type = file_data['mime_type']

                telegram_file = await context.bot.get_file(file_id)
                file_bytes = await telegram_file.download_as_bytearray()

                msg.add_attachment(
                    file_bytes,
                    maintype=mime_type.split('/')[0],
                    subtype=mime_type.split('/')[1],
                    filename=f"Позиция_{pos_index}_{file_name}",
                )
                logger.info(f"Дополнительный файл '{file_name}' для позиции {pos_index} прикреплен к письму")
            except Exception as e:
                logger.error(f"Ошибка при скачивании или прикреплении файла '{file_name}' для позиции {pos_index}: {e}")
                msg.set_content(msg.get_content() + f"\n\nВнимание: Не удалось прикрепить файл '{file_name}' для позиции {pos_index} из-за ошибки: {e}")


    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_LOGIN, EMAIL_PASSWORD)
            server.send_message(msg)
        logger.info(f"Письмо успешно отправлено на {to_email} с копией {cc_email}")
        return True
    except Exception as e:
        logger.error(f"Ошибка при отправке письма: {e}")
        raise

# === Telegram Handlers ===

async def initial_message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает первое сообщение от пользователя (или когда диалог неактивен)
    и предлагает кнопку "Создать заявку".
    """
    keyboard = [[KeyboardButton("Создать заявку")]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=False, resize_keyboard=True)
    await update.message.reply_text(
        "Привет! Я бот для создания заявок. Нажмите кнопку, чтобы начать.",
        reply_markup=reply_markup
    )

async def start_conversation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Начинает новый разговор (после нажатия кнопки "Создать заявку")
    и предлагает пользователю выбрать проект.
    """
    chat_id = update.effective_chat.id
    user = update.effective_user

    first_name = user.first_name if user.first_name else ""
    last_name = user.last_name if user.last_name else ""
    user_full_name = f"{first_name} {last_name}".strip()
    telegram_id_or_username = user.username if user.username else str(user.id)

    user_state[chat_id] = {
        "user_full_name": user_full_name,
        "telegram_id_or_username": telegram_id_or_username,
        "project": None,
        "object": None,
        "positions": [],
    }
    logger.info(f"User {user_full_name} ({telegram_id_or_username}) started conversation.")

    await update.message.reply_text("Начинаем создание заявки...", reply_markup=ReplyKeyboardRemove())

    keyboard = [[InlineKeyboardButton(p, callback_data=p)] for p in projects]
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите проект:", reply_markup=reply_markup)
    return PROJECT

async def project_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор проекта и предлагает выбрать объект."""
    query = update.callback_query
    await query.answer()

    user_state[query.message.chat.id]["project"] = query.data
    logger.info(f"Chat {query.message.chat.id}: Project selected - {query.data}")

    keyboard = [[InlineKeyboardButton(o, callback_data=o)] for o in objects]
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text("Выберите объект:", reply_markup=reply_markup)
    return OBJECT

async def object_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор объекта и запрашивает наименование позиции."""
    query = update.callback_query
    await query.answer()

    user_state[query.message.chat.id]["object"] = query.data
    logger.info(f"Chat {query.message.chat.id}: Object selected - {query.data}")
    await query.edit_message_text("Введите наименование позиции:")
    return NAME

async def name_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает наименование позиции и предлагает выбрать единицу измерения."""
    user_state[update.effective_chat.id]["current"] = {"name": update.message.text, "file_data": []}
    logger.info(f"Chat {update.effective_chat.id}: Position name entered - {update.message.text}")

    keyboard = [[InlineKeyboardButton(u, callback_data=u)] for u in units]
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите единицу измерения:", reply_markup=reply_markup)
    return UNIT

async def unit_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор единицы измерения и запрашивает количество."""
    query = update.callback_query
    await query.answer()

    user_state[query.message.chat.id]["current"]["unit"] = query.data
    logger.info(f"Chat {query.message.chat.id}: Unit selected - {query.data}")
    await query.edit_message_text("Введите количество:")
    return QUANTITY

async def quantity_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает количество и предлагает выбрать модуль. Добавлена базовая валидация."""
    chat_id = update.effective_chat.id
    try:
        quantity = float(update.message.text)
        user_state[chat_id]["current"]["quantity"] = quantity
        logger.info(f"Chat {chat_id}: Quantity entered - {quantity}")
    except ValueError:
        logger.warning(f"Chat {chat_id}: Invalid quantity format - '{update.message.text}'")
        await update.message.reply_text("Неверный формат количества. Пожалуйста, введите число (например, 5 или 3.5):")
        return QUANTITY

    buttons_per_row = 5
    keyboard = []
    current_row = []
    for i, m in enumerate(modules):
        current_row.append(InlineKeyboardButton(m, callback_data=m))
        if (i + 1) % buttons_per_row == 0 or (i + 1) == len(modules):
            keyboard.append(current_row)
            current_row = []
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("К какому модулю относится позиция?", reply_markup=reply_markup)
    return MODULE

async def module_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает выбор модуля и переходит к выбору даты поставки для текущей позиции.
    """
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    user_state[chat_id]["current"]["module"] = query.data
    logger.info(f"Chat {chat_id}: Module selected - {query.data}. Requesting delivery date for this position.")
    current_date = date.today()
    reply_markup = create_calendar_keyboard(current_date.year, current_date.month, prefix="POS_CAL_")
    await query.edit_message_text("Выберите желаемую дату поставки для этой позиции:", reply_markup=reply_markup)
    return POSITION_DELIVERY_DATE

# --- Calendar utility functions ---
def create_calendar_keyboard(year, month, prefix="CAL_"):
    """
    Создает инлайн-клавиатуру для выбора даты.
    """
    keyboard = []
    # Header: Month and Year
    keyboard.append([InlineKeyboardButton(f"{calendar.month_name[month]} {year}", callback_data=f"{prefix}IGNORE")])

    # Weekday headers
    week_days = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    keyboard.append([InlineKeyboardButton(day, callback_data=f"{prefix}IGNORE") for day in week_days])

    my_calendar = calendar.monthcalendar(year, month)
    for week in my_calendar:
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(" ", callback_data=f"{prefix}IGNORE"))
            else:
                date_str = f"{day:02d}.{month:02d}.{year}"
                row.append(InlineKeyboardButton(str(day), callback_data=f"{prefix}DATE_{date_str}"))
        keyboard.append(row)

    # Navigation buttons
    prev_month_year = (month - 1) if month != 1 else 12
    prev_year = year if month != 1 else year - 1
    next_month_year = (month + 1) if month != 12 else 1
    next_year = year if month != 12 else year + 1

    keyboard.append([
        InlineKeyboardButton("<<", callback_data=f"{prefix}NAV_{prev_year}_{prev_month_year}"),
        InlineKeyboardButton("Отмена", callback_data="cancel_dialog"),
        InlineKeyboardButton(">>", callback_data=f"{prefix}NAV_{next_year}_{next_month_year}")
    ])
    return InlineKeyboardMarkup(keyboard)


async def process_position_calendar_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает нажатия на кнопки календаря для выбора даты поставки отдельной позиции.
    """
    query = update.callback_query
    await query.answer()
    data = query.data
    chat_id = query.message.chat.id

    if data.startswith("POS_CAL_NAV_"):
        parts = data.split('_')
        year = int(parts[3])
        month = int(parts[4])
        if month > 12: month = 1; year += 1
        elif month < 1: month = 12; year -= 1
        reply_markup = create_calendar_keyboard(year, month, prefix="POS_CAL_")
        await query.edit_message_text("Выберите желаемую дату поставки для этой позиции:", reply_markup=reply_markup)
        return POSITION_DELIVERY_DATE
    elif data.startswith("POS_CAL_DATE_"):
        selected_date_str = data.replace("POS_CAL_DATE_", "")
        user_state[chat_id]["current"]["delivery_date"] = selected_date_str
        logger.info(f"Chat {chat_id}: Position delivery date selected - {selected_date_str}. Now asking about attachments.")
        keyboard = [
            [InlineKeyboardButton("Прикрепить файл", callback_data="attach_file")],
            [InlineKeyboardButton("Прикрепить ссылку", callback_data="attach_link")],
            [InlineKeyboardButton("Продолжить (без вложений)", callback_data="no_attachments")],
            [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Хотите прикрепить файл или ссылку к этой позиции?", reply_markup=reply_markup)
        return ATTACHMENT_CHOICE

    return POSITION_DELIVERY_DATE


async def attachment_choice_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает выбор пользователя по прикреплению файла/ссылки или продолжению.
    """
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    choice = query.data

    if choice == "attach_file":
        await query.edit_message_text("Пришлите файл (фото/документ):")
        return FILE_INPUT
    elif choice == "attach_link":
        await query.edit_message_text("Введите ссылку:")
        return LINK_INPUT
    elif choice == "no_attachments":
        # Если нет вложений, сразу добавляем текущую позицию и спрашиваем о новых
        user_state[chat_id]["positions"].append(user_state[chat_id].pop("current"))
        await ask_add_more_positions(update, context, query.message)
        return CONFIRM_ADD_MORE
    else:
        await query.edit_message_text("Неизвестный выбор. Пожалуйста, попробуйте снова.")
        return ATTACHMENT_CHOICE


async def file_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает прикрепленный файл."""
    chat_id = update.effective_chat.id
    if update.message.document:
        file_id = update.message.document.file_id
        file_name = update.message.document.file_name
        mime_type = update.message.document.mime_type
    elif update.message.photo:
        file_id = update.message.photo[-1].file_id # Get the highest resolution photo
        file_name = f"photo_{file_id}.jpg" # Or generate a more descriptive name
        mime_type = "image/jpeg"
    else:
        await update.message.reply_text("Пожалуйста, прикрепите файл (документ или фото).")
        return FILE_INPUT

    if "file_data" not in user_state[chat_id]["current"]:
        user_state[chat_id]["current"]["file_data"] = []
    user_state[chat_id]["current"]["file_data"].append({
        "file_id": file_id,
        "file_name": file_name,
        "mime_type": mime_type
    })
    logger.info(f"Chat {chat_id}: File '{file_name}' attached to current position.")

    keyboard = [
        [InlineKeyboardButton("Добавить еще файл", callback_data="attach_file")],
        [InlineKeyboardButton("Добавить ссылку", callback_data="attach_link")],
        [InlineKeyboardButton("Завершить вложения и добавить позицию", callback_data="finish_attachments")],
        [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Файл прикреплен. Хотите добавить еще вложения к этой позиции?", reply_markup=reply_markup)
    return ATTACHMENT_CHOICE


async def link_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает введенную ссылку."""
    chat_id = update.effective_chat.id
    link = update.message.text
    user_state[chat_id]["current"]["link"] = link
    logger.info(f"Chat {chat_id}: Link '{link}' added to current position.")

    keyboard = [
        [InlineKeyboardButton("Добавить файл", callback_data="attach_file")],
        [InlineKeyboardButton("Добавить еще ссылку", callback_data="attach_link")],
        [InlineKeyboardButton("Завершить вложения и добавить позицию", callback_data="finish_attachments")],
        [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Ссылка добавлена. Хотите добавить еще вложения к этой позиции?", reply_markup=reply_markup)
    return ATTACHMENT_CHOICE


async def finish_attachments_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает завершение добавления вложений и добавляет текущую позицию в список.
    """
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id

    user_state[chat_id]["positions"].append(user_state[chat_id].pop("current"))
    logger.info(f"Chat {chat_id}: Current position with attachments added to list. Total positions: {len(user_state[chat_id]['positions'])}")

    await ask_add_more_positions(update, context, query.message)
    return CONFIRM_ADD_MORE


async def ask_add_more_positions(update: Update, context: ContextTypes.DEFAULT_TYPE, message):
    """
    Спрашивает пользователя, хочет ли он добавить еще позиции или завершить заявку.
    """
    keyboard = [
        [InlineKeyboardButton("Добавить еще позицию", callback_data="add_more_yes")],
        [InlineKeyboardButton("Завершить заявку", callback_data="add_more_no")],
        [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await message.edit_text("Позиция добавлена. Хотите добавить еще позиции или завершить заявку?", reply_markup=reply_markup)


async def confirm_add_more_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает выбор пользователя по добавлению новых позиций или завершению.
    """
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    choice = query.data

    if choice == "add_more_yes":
        logger.info(f"Chat {chat_id}: User chose to add more positions.")
        await query.edit_message_text("Введите наименование следующей позиции:")
        return NAME
    elif choice == "add_more_no":
        logger.info(f"Chat {chat_id}: User chose to finalize the application.")
        return await final_confirmation_menu(update, context) # Переход к меню финального подтверждения

    return CONFIRM_ADD_MORE

async def final_confirmation_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Показывает итоговую информацию по заявке и предлагает подтвердить или изменить.
    """
    chat_id = update.effective_chat.id
    data = user_state[chat_id]
    project = data["project"]
    object_name = data["object"]
    positions = data["positions"]

    summary = f"**Итоговая заявка:**\n"
    summary += f"Проект: `{project}`\n"
    summary += f"Объект: `{object_name}`\n\n"
    summary += "**Позиции:**\n"

    for i, pos in enumerate(positions):
        summary += f"{i+1}. Наименование: `{pos.get('name', 'N/A')}`\n"
        summary += f"   Ед.изм.: `{pos.get('unit', 'N/A')}`\n"
        summary += f"   Количество: `{pos.get('quantity', 'N/A')}`\n"
        summary += f"   Модуль: `{pos.get('module', 'N/A')}`\n"
        summary += f"   Дата поставки: `{pos.get('delivery_date', 'Не указано')}`\n"
        if pos.get('link'):
            summary += f"   Ссылка: {pos['link']}\n"
        if pos.get('file_data'):
            file_names = ", ".join([f['file_name'] for f in pos['file_data']])
            summary += f"   Файлы: {file_names}\n"
        summary += "\n"

    keyboard = [
        [InlineKeyboardButton("Подтвердить и отправить", callback_data="final_confirm_send")],
        [InlineKeyboardButton("Изменить данные", callback_data="edit_application")],
        [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=chat_id,
        text=summary,
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )
    return FINAL_CONFIRMATION


async def final_confirm_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает финальное подтверждение и предлагает выбрать специалиста.
    """
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    choice = query.data

    if choice == "final_confirm_send":
        keyboard = [
            [InlineKeyboardButton("Алеся Забавская", callback_data="specialist_alesya")],
            [InlineKeyboardButton("Дмитрий Карп", callback_data="specialist_dmitry")],
            [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Кому направить заявку?", reply_markup=reply_markup)
        return SPECIALIST_SELECTION
    elif choice == "edit_application":
        await query.edit_message_text("Что хотите изменить?", reply_markup=create_edit_menu_keyboard())
        return EDIT_MENU
    else:
        await query.edit_message_text("Неизвестный выбор. Пожалуйста, попробуйте снова.")
        return FINAL_CONFIRMATION


async def specialist_selection_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает выбор специалиста и отправляет заявку.
    """
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    specialist_choice = query.data

    to_email = ""
    if specialist_choice == "specialist_alesya":
        to_email = SPECIALIST_EMAIL_ALESYA
        specialist_name = "Алесе Забавской"
    elif specialist_choice == "specialist_dmitry":
        to_email = SPECIALIST_EMAIL_DMITRY
        specialist_name = "Дмитрию Карпу"
    else:
        await query.edit_message_text("Неверный выбор специалиста. Пожалуйста, попробуйте снова.",
                                      reply_markup=InlineKeyboardMarkup([
                                          [InlineKeyboardButton("Алеся Забавская", callback_data="specialist_alesya")],
                                          [InlineKeyboardButton("Дмитрий Карп", callback_data="specialist_dmitry")],
                                          [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]
                                      ]))
        return SPECIALIST_SELECTION

    data = user_state[chat_id]
    project = data["project"]
    object_name = data["object"]
    positions = data["positions"]
    user_full_name = data["user_full_name"]
    telegram_id_or_username = data["telegram_id_or_username"]

    try:
        await send_email(chat_id, project, object_name, positions, user_full_name, telegram_id_or_username, to_email, CC_EMAIL, context)
        await query.edit_message_text(f"Заявка успешно отправлена {specialist_name}!")
        logger.info(f"Chat {chat_id}: Application successfully sent to {to_email} with CC {CC_EMAIL}.")
    except Exception as e:
        logger.error(f"Chat {chat_id}: Error sending email: {e}")
        await query.edit_message_text(f"Произошла ошибка при отправке заявки. Пожалуйста, попробуйте еще раз. Ошибка: {e}")

    # Очищаем состояние пользователя после завершения заявки
    if chat_id in user_state:
        del user_state[chat_id]
        logger.info(f"Chat {chat_id}: User state cleared.")

    return ConversationHandler.END


# --- Edit menu handlers ---
def create_edit_menu_keyboard():
    """Создает клавиатуру для меню редактирования."""
    keyboard = [
        [InlineKeyboardButton("Изменить проект", callback_data="edit_project")],
        [InlineKeyboardButton("Изменить объект", callback_data="edit_object")],
        [InlineKeyboardButton("Изменить общую дату поставки", callback_data="edit_global_delivery_date")],
        [InlineKeyboardButton("Изменить позицию", callback_data="edit_position_select")],
        [InlineKeyboardButton("Завершить редактирование", callback_data="finish_editing")],
        [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]
    ]
    return InlineKeyboardMarkup(keyboard)

async def edit_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор в меню редактирования."""
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    choice = query.data

    if choice == "edit_project":
        keyboard = [[InlineKeyboardButton(p, callback_data=f"edit_project_{p}")] for p in projects]
        keyboard.append([InlineKeyboardButton("Отмена редактирования", callback_data="cancel_edit")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Выберите новый проект:", reply_markup=reply_markup)
        return EDIT_FIELD_SELECTION
    elif choice == "edit_object":
        keyboard = [[InlineKeyboardButton(o, callback_data=f"edit_object_{o}")] for o in objects]
        keyboard.append([InlineKeyboardButton("Отмена редактирования", callback_data="cancel_edit")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Выберите новый объект:", reply_markup=reply_markup)
        return EDIT_FIELD_SELECTION
    elif choice == "edit_global_delivery_date":
        current_date = date.today()
        reply_markup = create_calendar_keyboard(current_date.year, current_date.month, prefix="EDIT_CAL_")
        await query.edit_message_text("Выберите новую общую дату поставки:", reply_markup=reply_markup)
        return GLOBAL_DELIVERY_DATE_SELECTION
    elif choice == "edit_position_select":
        if not user_state[chat_id]["positions"]:
            await query.edit_message_text("Нет позиций для редактирования. Выберите другое действие.", reply_markup=create_edit_menu_keyboard())
            return EDIT_MENU
        keyboard = [[InlineKeyboardButton(f"Позиция {i+1}: {pos['name']}", callback_data=f"select_pos_{i}")]
                    for i, pos in enumerate(user_state[chat_id]["positions"])]
        keyboard.append([InlineKeyboardButton("Отмена редактирования", callback_data="cancel_edit")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Выберите позицию для редактирования:", reply_markup=reply_markup)
        return SELECT_POSITION
    elif choice == "finish_editing":
        await query.edit_message_text("Редактирование завершено.")
        return await final_confirmation_menu(update, context)
    elif choice == "cancel_edit":
        await query.edit_message_text("Редактирование отменено.")
        return await final_confirmation_menu(update, context) # Возврат к финальному подтверждению
    return EDIT_MENU

async def process_edited_field_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор нового значения для проекта/объекта."""
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    data = query.data

    if data.startswith("edit_project_"):
        new_value = data.replace("edit_project_", "")
        user_state[chat_id]["project"] = new_value
        logger.info(f"Chat {chat_id}: Project changed to {new_value}.")
        await query.edit_message_text(f"Проект изменен на: {new_value}")
    elif data.startswith("edit_object_"):
        new_value = data.replace("edit_object_", "")
        user_state[chat_id]["object"] = new_value
        logger.info(f"Chat {chat_id}: Object changed to {new_value}.")
        await query.edit_message_text(f"Объект изменен на: {new_value}")
    else:
        await query.edit_message_text("Неизвестное изменение.")

    await query.message.reply_text("Что еще хотите изменить?", reply_markup=create_edit_menu_keyboard())
    return EDIT_MENU

async def process_global_calendar_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает нажатия на кнопки календаря для выбора общей даты поставки.
    """
    query = update.callback_query
    await query.answer()
    data = query.data
    chat_id = query.message.chat.id

    if data.startswith("EDIT_CAL_NAV_"):
        parts = data.split('_')
        year = int(parts[3])
        month = int(parts[4])
        if month > 12: month = 1; year += 1
        elif month < 1: month = 12; year -= 1
        reply_markup = create_calendar_keyboard(year, month, prefix="EDIT_CAL_")
        await query.edit_message_text("Выберите новую общую дату поставки:", reply_markup=reply_markup)
        return GLOBAL_DELIVERY_DATE_SELECTION
    elif data.startswith("EDIT_CAL_DATE_"):
        selected_date_str = data.replace("EDIT_CAL_DATE_", "")
        # Применяем общую дату ко всем позициям без даты поставки
        for pos in user_state[chat_id]["positions"]:
            if "delivery_date" not in pos or pos["delivery_date"] == "Не указано":
                pos["delivery_date"] = selected_date_str
        logger.info(f"Chat {chat_id}: Global delivery date set to {selected_date_str} for all relevant positions.")
        await query.edit_message_text(f"Общая дата поставки установлена: {selected_date_str}")
        await query.message.reply_text("Что еще хотите изменить?", reply_markup=create_edit_menu_keyboard())
        return EDIT_MENU
    return GLOBAL_DELIVERY_DATE_SELECTION

async def select_position_to_edit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор позиции для редактирования и предлагает поля для изменения."""
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    selected_index = int(query.data.replace("select_pos_", ""))
    user_state[chat_id]["editing_position_index"] = selected_index
    current_pos = user_state[chat_id]["positions"][selected_index]

    keyboard = [
        [InlineKeyboardButton("Наименование", callback_data="edit_field_name")],
        [InlineKeyboardButton("Единица измерения", callback_data="edit_field_unit")],
        [InlineKeyboardButton("Количество", callback_data="edit_field_quantity")],
        [InlineKeyboardButton("Модуль", callback_data="edit_field_module")],
        [InlineKeyboardButton("Дата поставки", callback_data="edit_field_delivery_date")],
        [InlineKeyboardButton("Ссылка", callback_data="edit_field_link")],
        [InlineKeyboardButton("Файлы (удалить/добавить)", callback_data="edit_field_files")],
        [InlineKeyboardButton("Отмена", callback_data="cancel_edit_field")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(f"Вы выбрали позицию: {current_pos['name']}. Что хотите изменить?", reply_markup=reply_markup)
    return EDIT_FIELD_SELECTION


async def edit_field_selection_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор поля для редактирования в позиции."""
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    field_choice = query.data

    user_state[chat_id]["editing_field"] = field_choice

    if field_choice == "edit_field_name":
        await query.edit_message_text("Введите новое наименование:")
        return EDIT_FIELD_INPUT
    elif field_choice == "edit_field_unit":
        keyboard = [[InlineKeyboardButton(u, callback_data=f"edit_unit_{u}")] for u in units]
        keyboard.append([InlineKeyboardButton("Отмена", callback_data="cancel_edit_field")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Выберите новую единицу измерения:", reply_markup=reply_markup)
        return EDITING_UNIT
    elif field_choice == "edit_field_quantity":
        await query.edit_message_text("Введите новое количество:")
        return EDIT_FIELD_INPUT
    elif field_choice == "edit_field_module":
        buttons_per_row = 5
        keyboard = []
        current_row = []
        for i, m in enumerate(modules):
            current_row.append(InlineKeyboardButton(m, callback_data=f"edit_module_{m}"))
            if (i + 1) % buttons_per_row == 0 or (i + 1) == len(modules):
                keyboard.append(current_row)
                current_row = []
        keyboard.append([InlineKeyboardButton("Отмена", callback_data="cancel_edit_field")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Выберите новый модуль:", reply_markup=reply_markup)
        return EDITING_MODULE
    elif field_choice == "edit_field_delivery_date":
        current_date = date.today()
        reply_markup = create_calendar_keyboard(current_date.year, current_date.month, prefix="POS_CAL_") # Используем тот же префикс для позиций
        await query.edit_message_text("Выберите новую дату поставки для этой позиции:", reply_markup=reply_markup)
        return POSITION_DELIVERY_DATE # Возвращаемся в состояние выбора даты для позиции
    elif field_choice == "edit_field_link":
        await query.edit_message_text("Введите новую ссылку (или 'нет' для удаления):")
        return EDIT_FIELD_INPUT
    elif field_choice == "edit_field_files":
        current_pos_index = user_state[chat_id]["editing_position_index"]
        current_pos = user_state[chat_id]["positions"][current_pos_index]
        file_options_keyboard = []
        if current_pos.get("file_data"):
            for i, file_item in enumerate(current_pos["file_data"]):
                file_options_keyboard.append([InlineKeyboardButton(f"Удалить: {file_item['file_name']}", callback_data=f"delete_file_{i}")])
        file_options_keyboard.append([InlineKeyboardButton("Добавить новый файл", callback_data="add_new_file")])
        file_options_keyboard.append([InlineKeyboardButton("Отмена", callback_data="cancel_edit_field")])
        reply_markup = InlineKeyboardMarkup(file_options_keyboard)
        await query.edit_message_text("Управление файлами:", reply_markup=reply_markup)
        return FILE_INPUT # Переиспользуем FILE_INPUT для обработки добавления/удаления файлов
    elif field_choice == "cancel_edit_field":
        await query.edit_message_text("Редактирование поля отменено. Что еще хотите изменить?", reply_markup=create_edit_menu_keyboard())
        return EDIT_MENU
    return EDIT_FIELD_SELECTION


async def edit_field_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает ввод нового значения для текстовых полей."""
    chat_id = update.effective_chat.id
    current_pos_index = user_state[chat_id]["editing_position_index"]
    field_to_edit = user_state[chat_id]["editing_field"]
    new_value = update.message.text

    if field_to_edit == "edit_field_name":
        user_state[chat_id]["positions"][current_pos_index]["name"] = new_value
        await update.message.reply_text(f"Наименование изменено на: {new_value}")
    elif field_to_edit == "edit_field_quantity":
        try:
            quantity = float(new_value)
            user_state[chat_id]["positions"][current_pos_index]["quantity"] = quantity
            await update.message.reply_text(f"Количество изменено на: {quantity}")
        except ValueError:
            await update.message.reply_text("Неверный формат количества. Пожалуйста, введите число.")
            return EDIT_FIELD_INPUT
    elif field_to_edit == "edit_field_link":
        if new_value.lower() == 'нет':
            user_state[chat_id]["positions"][current_pos_index].pop("link", None)
            await update.message.reply_text("Ссылка удалена.")
        else:
            user_state[chat_id]["positions"][current_pos_index]["link"] = new_value
            await update.message.reply_text(f"Ссылка изменена на: {new_value}")
    elif field_to_edit == "edit_field_files":
        # Handle file input for adding new files
        if update.message.document:
            file_id = update.message.document.file_id
            file_name = update.message.document.file_name
            mime_type = update.message.document.mime_type
        elif update.message.photo:
            file_id = update.message.photo[-1].file_id
            file_name = f"photo_{file_id}.jpg"
            mime_type = "image/jpeg"
        else:
            await update.message.reply_text("Пожалуйста, прикрепите файл (документ или фото).")
            return FILE_INPUT

        if "file_data" not in user_state[chat_id]["positions"][current_pos_index]:
            user_state[chat_id]["positions"][current_pos_index]["file_data"] = []
        user_state[chat_id]["positions"][current_pos_index]["file_data"].append({
            "file_id": file_id,
            "file_name": file_name,
            "mime_type": mime_type
        })
        await update.message.reply_text(f"Файл '{file_name}' добавлен.")
        # After adding, offer options to add more/delete or go back to edit menu
        await query_for_more_file_actions(update, context, current_pos_index)
        return FILE_INPUT # Stay in file input mode for more actions

    await update.message.reply_text("Что еще хотите изменить в этой позиции?", reply_markup=get_edit_position_fields_keyboard())
    return EDIT_FIELD_SELECTION

async def query_for_more_file_actions(update: Update, context: ContextTypes.DEFAULT_TYPE, current_pos_index):
    """Helper to display file management options after a file action."""
    chat_id = update.effective_chat.id
    current_pos = user_state[chat_id]["positions"][current_pos_index]
    file_options_keyboard = []
    if current_pos.get("file_data"):
        for i, file_item in enumerate(current_pos["file_data"]):
            file_options_keyboard.append([InlineKeyboardButton(f"Удалить: {file_item['file_name']}", callback_data=f"delete_file_{i}")])
    file_options_keyboard.append([InlineKeyboardButton("Добавить новый файл", callback_data="add_new_file")])
    file_options_keyboard.append([InlineKeyboardButton("Завершить редактирование файлов", callback_data="finish_file_editing")])
    reply_markup = InlineKeyboardMarkup(file_options_keyboard)
    await context.bot.send_message(chat_id=chat_id, text="Что еще хотите сделать с файлами этой позиции?", reply_markup=reply_markup)


async def process_edited_unit_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор новой единицы измерения при редактировании позиции."""
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    current_pos_index = user_state[chat_id]["editing_position_index"]
    new_unit = query.data.replace("edit_unit_", "")

    user_state[chat_id]["positions"][current_pos_index]["unit"] = new_unit
    await query.edit_message_text(f"Единица измерения изменена на: {new_unit}")
    await query.message.reply_text("Что еще хотите изменить в этой позиции?", reply_markup=get_edit_position_fields_keyboard())
    return EDIT_FIELD_SELECTION

async def process_edited_module_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор нового модуля при редактировании позиции."""
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    current_pos_index = user_state[chat_id]["editing_position_index"]
    new_module = query.data.replace("edit_module_", "")

    user_state[chat_id]["positions"][current_pos_index]["module"] = new_module
    await query.edit_message_text(f"Модуль изменен на: {new_module}")
    await query.message.reply_text("Что еще хотите изменить в этой позиции?", reply_markup=get_edit_position_fields_keyboard())
    return EDIT_FIELD_SELECTION


async def process_edited_file_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает действия с файлами (удаление, добавление)."""
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    current_pos_index = user_state[chat_id]["editing_position_index"]
    current_pos = user_state[chat_id]["positions"][current_pos_index]
    action = query.data

    if action.startswith("delete_file_"):
        file_index = int(action.replace("delete_file_", ""))
        if "file_data" in current_pos and len(current_pos["file_data"]) > file_index:
            deleted_file_name = current_pos["file_data"].pop(file_index)["file_name"]
            await query.edit_message_text(f"Файл '{deleted_file_name}' удален.")
        else:
            await query.edit_message_text("Не удалось удалить файл.")
        await query_for_more_file_actions(update, context, current_pos_index)
        return FILE_INPUT
    elif action == "add_new_file":
        await query.edit_message_text("Пришлите новый файл:")
        return FILE_INPUT # Expecting a file message next
    elif action == "finish_file_editing":
        await query.edit_message_text("Редактирование файлов завершено. Что еще хотите изменить в этой позиции?", reply_markup=get_edit_position_fields_keyboard())
        return EDIT_FIELD_SELECTION
    return FILE_INPUT # Stay in FILE_INPUT state for file related actions


def get_edit_position_fields_keyboard():
    """Возвращает клавиатуру для выбора поля для редактирования в позиции."""
    keyboard = [
        [InlineKeyboardButton("Наименование", callback_data="edit_field_name")],
        [InlineKeyboardButton("Единица измерения", callback_data="edit_field_unit")],
        [InlineKeyboardButton("Количество", callback_data="edit_field_quantity")],
        [InlineKeyboardButton("Модуль", callback_data="edit_field_module")],
        [InlineKeyboardButton("Дата поставки", callback_data="edit_field_delivery_date")],
        [InlineKeyboardButton("Ссылка", callback_data="edit_field_link")],
        [InlineKeyboardButton("Файлы (удалить/добавить)", callback_data="edit_field_files")],
        [InlineKeyboardButton("Завершить редактирование позиции", callback_data="finish_edit_position")],
        [InlineKeyboardButton("Отмена", callback_data="cancel_edit_field")]
    ]
    return InlineKeyboardMarkup(keyboard)

async def finish_edit_position_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает завершение редактирования текущей позиции."""
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("Редактирование позиции завершено. Что еще хотите изменить в заявке?", reply_markup=create_edit_menu_keyboard())
    if "editing_position_index" in user_state[query.message.chat.id]:
        del user_state[query.message.chat.id]["editing_position_index"]
    if "editing_field" in user_state[query.message.chat.id]:
        del user_state[query.message.chat.id]["editing_field"]
    return EDIT_MENU

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отменяет текущий диалог."""
    chat_id = update.effective_chat.id
    if chat_id in user_state:
        del user_state[chat_id]
        logger.info(f"Chat {chat_id}: Dialog canceled and user state cleared.")

    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(
            "Заявка отменена. Чтобы начать новую, нажмите 'Создать заявку'.",
            reply_markup=ReplyKeyboardMarkup([["Создать заявку"]], one_time_keyboard=False, resize_keyboard=True)
        )
    else:
        await update.message.reply_text(
            "Заявка отменена. Чтобы начать новую, нажмите 'Создать заявку'.",
            reply_markup=ReplyKeyboardMarkup([["Создать заявку"]], one_time_keyboard=False, resize_keyboard=True)
        )
    return ConversationHandler.END

async def unknown(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает неизвестные команды или сообщения."""
    await update.message.reply_text("Извините, я не понял вашу команду или сообщение. Пожалуйста, используйте кнопки или начните новую заявку.")
    return ConversationHandler.END


async def main():
    """Запускает бота."""
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex("^Создать заявку$"), start_conversation),
            CommandHandler("start", start_conversation)
        ],
        states={
            PROJECT: [CallbackQueryHandler(project_handler, pattern="^(?!cancel_dialog$).*")],
            OBJECT: [CallbackQueryHandler(object_handler, pattern="^(?!cancel_dialog$).*")],
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, name_handler)],
            UNIT: [CallbackQueryHandler(unit_handler, pattern="^(?!cancel_dialog$).*")],
            QUANTITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, quantity_handler)],
            MODULE: [CallbackQueryHandler(module_handler, pattern="^(?!cancel_dialog$).*")],
            POSITION_DELIVERY_DATE: [
                CallbackQueryHandler(process_position_calendar_callback, pattern="^POS_CAL_"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            ATTACHMENT_CHOICE: [
                CallbackQueryHandler(attachment_choice_handler, pattern="^(attach_file|attach_link|no_attachments|finish_attachments)$")
            ],
            FILE_INPUT: [
                MessageHandler(filters.Document.ALL | filters.PHOTO, file_input_handler),
                CallbackQueryHandler(process_edited_file_action, pattern="^(delete_file_|add_new_file|finish_file_editing)$"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            LINK_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, link_input_handler)],
            CONFIRM_ADD_MORE: [
                CallbackQueryHandler(confirm_add_more_handler, pattern="^(add_more_yes|add_more_no)$")
            ],
            EDIT_MENU: [
                CallbackQueryHandler(edit_menu_handler, pattern="^(edit_project|edit_object|edit_global_delivery_date|edit_position_select|finish_editing|cancel_edit)$")
            ],
            SELECT_POSITION: [
                CallbackQueryHandler(select_position_to_edit, pattern="^select_pos_"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            EDIT_FIELD_SELECTION: [
                CallbackQueryHandler(edit_field_selection_handler, pattern="^(edit_field_name|edit_field_unit|edit_field_quantity|edit_field_module|edit_field_delivery_date|edit_field_link|edit_field_files|cancel_edit_field|finish_edit_position)$")
            ],
            EDIT_FIELD_INPUT: [
                MessageHandler(filters.TEXT | filters.Document.ALL | filters.PHOTO | filters.TEXT & ~filters.COMMAND, edit_field_input_handler),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            EDITING_UNIT: [
                CallbackQueryHandler(process_edited_unit_selection, pattern="^edit_unit_"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            EDITING_MODULE: [
                CallbackQueryHandler(process_edited_module_selection, pattern="^edit_module_"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            GLOBAL_DELIVERY_DATE_SELECTION: [
                CallbackQueryHandler(process_global_calendar_callback, pattern="^(CAL_|EDIT_CAL_)\\d+_\\d+?$"), # Уточнили паттерн
                CallbackQueryHandler(process_global_calendar_callback, pattern="^(CAL_|EDIT_CAL_DATE_)"), # Уточнили паттерн для выбора даты
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            FINAL_CONFIRMATION: [
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$"),
                CallbackQueryHandler(final_confirm_handler)
            ],
            SPECIALIST_SELECTION: [
                CallbackQueryHandler(specialist_selection_handler, pattern="^specialist_"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CallbackQueryHandler(cancel, pattern="^cancel_dialog$"),
            MessageHandler(filters.COMMAND | filters.TEXT, unknown)
        ],
    )

    app.add_handler(conv_handler)

    app.add_handler(CommandHandler("start", initial_message_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, initial_message_handler))

    await app.run_polling()

if __name__ == '__main__':
    nest_asyncio.apply() # Apply nest_asyncio to allow running asyncio loops in non-async environments

    loop = asyncio.get_event_loop()
    try:
        loop.run_until_complete(main())
    except RuntimeError as e:
        if "cannot close a running event loop" in str(e):
            # This handles cases where the event loop might already be running,
            # which can happen in certain interactive environments.
            logger.warning(f"RuntimeError detected: {e}. Attempting to run main() directly.")
            loop.run_until_complete(main())
        else:
            raise e
    except Exception as e:
        logger.error(f"An unexpected error occurred: {e}")
        raise e